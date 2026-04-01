# Builds the Excel reconciliation report.
# Creates a workbook with 7 sheets covering summary KPIs, matched pairs,
# exceptions with DeepSeek investigation notes, and full transaction lists.
# Uses openpyxl for writing and xlsxwriter-style formatting via openpyxl styles.

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colour constants for header rows
HEADER_FILL_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="1E6B3C", end_color="1E6B3C", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="7B1E1E", end_color="7B1E1E", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="7B4B1E", end_color="7B4B1E", fill_type="solid")
HEADER_FILL_GREY = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

SEVERITY_COLOURS = {
    "High": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Medium": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Low": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}


def build_excel_report(state: dict, report_path: str) -> None:
    # Creates and saves the full Excel report to report_path.
    # All 7 sheets are written from the pipeline state.

    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    write_summary_sheet(wb, state)
    write_matched_sheet(wb, state)
    write_exceptions_sheet(wb, state)
    write_bank_only_sheet(wb, state)
    write_ledger_only_sheet(wb, state)
    write_all_transactions_sheet(wb, state)
    write_all_ledger_sheet(wb, state)

    wb.save(report_path)


# --- Sheet writers ---

def write_summary_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("Summary")

    total_bank = state.get("total_bank", 0)
    matched_count = state.get("matched_count", 0)
    exceptions = state.get("exceptions", [])
    exception_count = len(exceptions)
    high_risk_count = sum(1 for e in exceptions if e.get("severity") == "High")

    match_rate = (matched_count / total_bank * 100) if total_bank > 0 else 0

    unmatched_bank = [
        t for t in state.get("bank_transactions", [])
        if not t.get("matched")
    ]
    unmatched_amount = sum(abs(t.get("amount", 0)) for t in unmatched_bank)

    rows = [
        ("OpenReco — Reconciliation Report", ""),
        ("", ""),
        ("Period Start", state.get("period_start", "")),
        ("Period End", state.get("period_end", "")),
        ("Bank File", os.path.basename(state.get("bank_file_path", ""))),
        ("Ledger File", os.path.basename(state.get("ledger_file_path", ""))),
        ("Report Generated", date.today().isoformat()),
        ("", ""),
        ("Total Bank Transactions", total_bank),
        ("Total Ledger Entries", state.get("total_ledger", 0)),
        ("Matched", matched_count),
        ("Match Rate", f"{match_rate:.1f}%"),
        ("Total Exceptions", exception_count),
        ("High Risk Exceptions", high_risk_count),
        ("Total Unmatched Amount", f"RM {unmatched_amount:,.2f}"),
        ("", ""),
        ("Narrative Summary", state.get("summary", "")),
    ]

    for row_index, (label, value) in enumerate(rows, start=1):
        label_cell = ws.cell(row=row_index, column=1, value=label)
        value_cell = ws.cell(row=row_index, column=2, value=value)

        if row_index == 1:
            label_cell.font = Font(bold=True, size=14)
        elif label in ("Matched", "Match Rate", "Total Exceptions", "High Risk Exceptions"):
            label_cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def write_matched_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("Matched")

    headers = [
        "Match ID", "Bank Date", "Bank Description", "Bank Amount",
        "Ledger Date", "Ledger Description", "Ledger Amount",
        "Match Type", "Confidence", "Reasoning",
    ]
    write_header_row(ws, headers, HEADER_FILL_GREEN)

    # Build lookup maps for fast access by ID
    bank_by_id = {t["id"]: t for t in state.get("bank_transactions", [])}
    ledger_by_id = {e["id"]: e for e in state.get("ledger_entries", [])}

    for match in state.get("matches", []):
        bank_txn = bank_by_id.get(match["bank_txn_id"], {})
        ledger_entry = ledger_by_id.get(match["ledger_entry_id"], {})

        ws.append([
            match.get("match_id", ""),
            bank_txn.get("date", ""),
            bank_txn.get("description", ""),
            abs(bank_txn.get("amount", 0)),
            ledger_entry.get("date", ""),
            ledger_entry.get("description", ""),
            abs(ledger_entry.get("amount", 0)),
            match.get("match_type", ""),
            f"{match.get('confidence', 0):.0%}",
            match.get("reasoning", ""),
        ])

    auto_width(ws, headers)


def write_exceptions_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("Exceptions")

    headers = [
        "Exception ID", "Type", "Source", "Amount", "Description",
        "Severity", "Investigation", "Recommended Action", "Suggested Match",
    ]
    write_header_row(ws, headers, HEADER_FILL_RED)

    for exc in state.get("exceptions", []):
        row_data = [
            exc.get("exception_id", ""),
            exc.get("type", ""),
            exc.get("item_source", ""),
            abs(exc.get("amount", 0)),
            exc.get("description", ""),
            exc.get("severity", ""),
            exc.get("investigation", ""),
            exc.get("resolution", ""),
            exc.get("suggested_match", ""),
        ]
        ws.append(row_data)

        # Colour the whole row based on severity
        severity = exc.get("severity", "")
        fill = SEVERITY_COLOURS.get(severity)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    auto_width(ws, headers)


def write_bank_only_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("Bank Only")

    headers = ["ID", "Date", "Description", "Reference", "Debit", "Credit", "Amount"]
    write_header_row(ws, headers, HEADER_FILL_ORANGE)

    for txn in state.get("bank_transactions", []):
        if not txn.get("matched"):
            ws.append([
                txn.get("id", ""),
                txn.get("date", ""),
                txn.get("description", ""),
                txn.get("reference", ""),
                txn.get("debit", 0),
                txn.get("credit", 0),
                abs(txn.get("amount", 0)),
            ])

    auto_width(ws, headers)


def write_ledger_only_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("Ledger Only")

    headers = ["ID", "Date", "Description", "Reference", "Amount", "Type"]
    write_header_row(ws, headers, HEADER_FILL_ORANGE)

    for entry in state.get("ledger_entries", []):
        if not entry.get("matched"):
            ws.append([
                entry.get("id", ""),
                entry.get("date", ""),
                entry.get("description", ""),
                entry.get("reference", ""),
                abs(entry.get("amount", 0)),
                entry.get("entry_type", ""),
            ])

    auto_width(ws, headers)


def write_all_transactions_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("All Transactions")

    headers = ["ID", "Date", "Description", "Reference", "Debit", "Credit", "Amount", "Matched", "Match ID", "Confidence"]
    write_header_row(ws, headers, HEADER_FILL_GREY)

    for txn in state.get("bank_transactions", []):
        ws.append([
            txn.get("id", ""),
            txn.get("date", ""),
            txn.get("description", ""),
            txn.get("reference", ""),
            txn.get("debit", 0),
            txn.get("credit", 0),
            abs(txn.get("amount", 0)),
            "Yes" if txn.get("matched") else "No",
            txn.get("match_id", ""),
            f"{txn.get('confidence', 0):.0%}" if txn.get("confidence") else "",
        ])

    auto_width(ws, headers)


def write_all_ledger_sheet(wb: Workbook, state: dict) -> None:
    ws = wb.create_sheet("All Ledger")

    headers = ["ID", "Date", "Description", "Reference", "Amount", "Type", "Matched", "Match ID"]
    write_header_row(ws, headers, HEADER_FILL_GREY)

    for entry in state.get("ledger_entries", []):
        ws.append([
            entry.get("id", ""),
            entry.get("date", ""),
            entry.get("description", ""),
            entry.get("reference", ""),
            abs(entry.get("amount", 0)),
            entry.get("entry_type", ""),
            "Yes" if entry.get("matched") else "No",
            entry.get("match_id", ""),
        ])

    auto_width(ws, headers)


# --- Helpers ---

def write_header_row(ws, headers: list, fill: PatternFill) -> None:
    # Writes a styled header row to the given worksheet.
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws, headers: list) -> None:
    # Sets each column width to roughly fit the header text.
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 14)

    # Make description and reasoning columns wider
    for i, header in enumerate(headers, start=1):
        if any(keyword in header.lower() for keyword in ["description", "reasoning", "investigation", "action", "summary"]):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 40
